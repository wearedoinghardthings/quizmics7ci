"""
app.py — QuizAgent CI
"""
import streamlit as st
import pandas as pd
import json, time, io, re, requests
from datetime import datetime, timedelta

import config, hashlib
from functools import lru_cache
from database import (
    init_db, search_agents, get_all_agents, upsert_agents,
    set_agent_published, publish_all_agents, unpublish_all_agents,
    get_agent_by_id, delete_agent, add_agent_manual,
    get_quiz_by_code, get_all_quizzes, get_quiz, create_quiz, update_quiz, delete_quiz,
    get_questions, add_question, delete_question, reorder_questions,
    create_session, save_quiz_progress, get_incomplete_session,
    submit_session, get_results, session_already_completed, get_stats,
    get_question_stats,
    get_surveillance,
    get_session_answers,
)

import re as _re

def _fmt(t):
    """Formate : **gras** *italique* [rouge/bleu/vert/orange]...[/] newline=<br>"""
    t = str(t)
    t = _re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', t)
    t = _re.sub(r'\*([^*]+?)\*',  r'<i>\1</i>', t)
    t = _re.sub(r'\[rouge\](.+?)\[/rouge\]',   r'<b style="color:#DC2626">\1</b>', t)
    t = _re.sub(r'\[bleu\](.+?)\[/bleu\]',     r'<b style="color:#1D4ED8">\1</b>', t)
    t = _re.sub(r'\[vert\](.+?)\[/vert\]',     r'<b style="color:#059669">\1</b>', t)
    t = _re.sub(r'\[orange\](.+?)\[/orange\]', r'<b style="color:#D97706">\1</b>', t)
    t = t.replace('\n', '<br>')
    return t

st.set_page_config(page_title=config.APP_TITLE, page_icon="📝",
                   layout="centered", initial_sidebar_state="collapsed")

# ═══════════════════════════════════════════════════════════════
#  CSS + DESIGN SYSTEM
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800;900&display=swap" rel="stylesheet">
<style>
*,*::before,*::after{box-sizing:border-box}
html,body,[class*="css"],input,textarea,button,select,.stMarkdown p,label{font-family:'Outfit',sans-serif!important}
#MainMenu,footer,header,.stDeployButton,[data-testid="stToolbar"],[data-testid="stDecoration"],[data-testid="stStatusWidget"],.stAppDeployButton{display:none!important}
.main,.main>div,section.main,.appview-container,.block-container{overflow:visible!important}
.block-container{padding:0 1rem 5rem!important;max-width:640px!important;margin:0 auto!important}
:root{
  --B:#1D4ED8;--Bd:#1E3A8A;--Bl:#EFF6FF;--Bm:#DBEAFE;
  --G:#059669;--O:#D97706;--R:#DC2626;--P:#7C3AED;
  --bg:#F8FAFF;--sur:#FFFFFF;--bor:#E2E8F4;--txt:#0F172A;--mu:#64748B;
  --r:12px;--sh:0 1px 3px rgba(15,23,42,.07),0 4px 12px rgba(15,23,42,.05);
  --shb:0 4px 20px rgba(29,78,216,.18)
}
/* ── Topbar ── */
.topbar{background:var(--Bd);margin:0 -1rem 1.5rem -1rem;padding:14px 20px;display:flex;align-items:center;gap:10px}
.tt{color:#fff;font-size:1.05rem;font-weight:700}
/* ── Hero ── */
.hero{background:linear-gradient(145deg,var(--Bd),var(--B));border-radius:18px;padding:36px 24px 32px;text-align:center;margin-bottom:20px;box-shadow:var(--shb)}
.hi{font-size:2.8rem;margin-bottom:6px}.ht{color:#fff;font-size:1.85rem;font-weight:900;margin:0 0 5px;letter-spacing:-.5px}.hs{color:rgba(255,255,255,.72);font-size:.95rem;margin:0}
/* ── Boutons ── */
.stButton>button{font-family:'Outfit',sans-serif!important;font-weight:600!important;font-size:15px!important;border-radius:10px!important;padding:13px 20px!important;width:100%!important;transition:all .15s!important;min-height:48px!important}
.stButton>button[kind="primary"]{background:var(--B)!important;border:none!important;color:#fff!important;box-shadow:0 3px 10px rgba(29,78,216,.3)!important}
.stButton>button[kind="primary"]:hover{background:var(--Bd)!important;transform:translateY(-1px)!important}
.stButton>button:not([kind="primary"]){background:var(--sur)!important;border:1.5px solid var(--bor)!important;color:var(--txt)!important}
/* ── Inputs ── */
.stTextInput>div>div>input,.stTextArea>div>div>textarea,.stNumberInput>div>div>input{font-family:'Outfit',sans-serif!important;font-size:16px!important;border-radius:10px!important;border:1.5px solid var(--bor)!important;padding:12px 14px!important}
.stTextInput>div>div>input:focus,.stTextArea>div>div>textarea:focus{border-color:var(--B)!important;box-shadow:0 0 0 3px rgba(29,78,216,.1)!important}
.stTextInput label,.stTextArea label,.stSelectbox label,.stNumberInput label{font-size:11px!important;font-weight:700!important;color:var(--mu)!important;text-transform:uppercase!important;letter-spacing:.08em!important}
/* ── Timer fixe ── */
#quiz-timer-bar{position:fixed!important;top:0!important;left:0!important;right:0!important;z-index:99999!important;background:rgba(248,250,255,.97)!important;backdrop-filter:blur(12px)!important;-webkit-backdrop-filter:blur(12px)!important;border-bottom:1px solid var(--bor)!important;box-shadow:0 2px 12px rgba(15,23,42,.08)!important;padding:8px 1rem 6px!important}
#quiz-timer-bar .inner{max-width:620px;margin:0 auto}
.tbox{border-radius:12px;padding:12px 20px;text-align:center;display:flex;align-items:center;justify-content:center;gap:14px}
.tn{background:linear-gradient(135deg,var(--Bd),var(--B));box-shadow:0 3px 12px rgba(29,78,216,.25)}
.tw{background:linear-gradient(135deg,#92400E,var(--O));box-shadow:0 3px 12px rgba(217,119,6,.25)}
.td{background:linear-gradient(135deg,#7F1D1D,var(--R));box-shadow:0 3px 12px rgba(220,38,38,.3);animation:pulse .7s infinite}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:.82}}
.ttime{color:#fff;font-size:2.2rem;font-weight:900;letter-spacing:4px;margin:0;line-height:1}
.tlbl{color:rgba(255,255,255,.75);font-size:.75rem;font-weight:600;margin:3px 0 0;text-transform:uppercase;letter-spacing:.08em}
.ticon{font-size:1.6rem}
/* ── Progression ── */
.prog{margin:6px 0 14px}.prog-row{display:flex;justify-content:space-between;font-size:.82rem;color:var(--mu);font-weight:500;margin-bottom:5px}
.prog-bg{background:var(--Bm);border-radius:99px;height:6px}.prog-fill{background:var(--B);height:6px;border-radius:99px;transition:width .3s}
/* ── Question ── */
.qcard{background:var(--sur);border:1.5px solid var(--bor);border-left:4px solid var(--B);border-radius:0 var(--r) var(--r) 0;padding:14px 16px 4px;box-shadow:var(--sh);margin:12px 0 0}
.qmeta{display:flex;align-items:center;gap:7px;margin-bottom:8px;flex-wrap:wrap}
.qnum{background:var(--Bl);color:var(--B);font-size:.72rem;font-weight:800;padding:2px 8px;border-radius:99px}
.qtype{color:var(--mu);font-size:.75rem;font-weight:500}.qpts{background:#F1F5F9;color:var(--mu);font-size:.72rem;font-weight:700;padding:2px 8px;border-radius:99px;margin-left:auto}
.qtxt{font-size:15px;font-weight:400;color:var(--txt);margin:0 0 10px;line-height:1.5}
.qtxt b{font-weight:800}
.qtxt [style*="color:#DC2626"]{color:#DC2626!important}
.qtxt [style*="color:#1D4ED8"]{color:#1D4ED8!important}
.qtxt [style*="color:#059669"]{color:#059669!important}
.qtxt [style*="color:#D97706"]{color:#D97706!important}
/* ── Radio/Checkbox ── */
.stRadio>div{gap:5px!important}
.stRadio>div>label{background:var(--bg)!important;border:1.5px solid var(--bor)!important;border-radius:10px!important;padding:11px 14px!important;font-size:14.5px!important;font-weight:500!important;min-height:46px!important;cursor:pointer!important;transition:all .12s!important}
.stRadio>div>label:hover{border-color:var(--B)!important;background:var(--Bl)!important}
div[data-testid="stCheckbox"] label{background:var(--bg);border:1.5px solid var(--bor);border-radius:10px;padding:11px 14px;font-size:14.5px!important;font-weight:500!important;min-height:46px;margin-bottom:5px;transition:all .12s}
div[data-testid="stCheckbox"] label:hover{border-color:var(--B);background:var(--Bl)}
/* ── Score ── */
.scard{background:linear-gradient(145deg,var(--Bd),var(--B));border-radius:18px;padding:32px 24px;text-align:center;box-shadow:var(--shb);margin:8px 0 20px}
.se{font-size:3rem;margin-bottom:4px}.sn{color:rgba(255,255,255,.8);font-size:.95rem;margin:0 0 6px}
.sv{color:#fff;font-size:2.8rem;font-weight:900;margin:0;letter-spacing:-1px}
.sp{color:rgba(255,255,255,.9);font-size:1.5rem;font-weight:700;margin:2px 0 6px}
.sq{color:rgba(255,255,255,.6);font-size:.85rem;margin:0}
.subcard{background:var(--Bl);border:2px solid var(--Bm);border-radius:16px;padding:30px 24px;text-align:center;margin:8px 0 20px}
/* ── Cards ── */
.card{background:var(--sur);border:1.5px solid var(--bor);border-radius:var(--r);padding:16px 18px;margin:8px 0;box-shadow:var(--sh)}
.slbl{font-size:.7rem;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:.1em;margin:20px 0 8px}
/* ── KPI cards ── */
.kpi-grid{display:grid;grid-template-columns:repeat(2,1fr);gap:10px;margin:0 0 20px}
.kpi{border-radius:14px;padding:18px 16px;box-shadow:var(--sh);position:relative;overflow:hidden}
.kpi::before{content:'';position:absolute;top:-20px;right:-20px;width:80px;height:80px;border-radius:50%;opacity:.12}
.kpi-b{background:linear-gradient(135deg,var(--Bl),#fff);border:1.5px solid var(--Bm)}.kpi-b::before{background:var(--B)}
.kpi-g{background:linear-gradient(135deg,#D1FAE5,#fff);border:1.5px solid #A7F3D0}.kpi-g::before{background:var(--G)}
.kpi-o{background:linear-gradient(135deg,#FEF3C7,#fff);border:1.5px solid #FDE68A}.kpi-o::before{background:var(--O)}
.kpi-p{background:linear-gradient(135deg,#EDE9FE,#fff);border:1.5px solid #DDD6FE}.kpi-p::before{background:var(--P)}
.kpi-icon{font-size:1.6rem;margin-bottom:8px}
.kpi-val{font-size:2rem;font-weight:900;margin:0;line-height:1}
.kpi-b .kpi-val{color:var(--B)}.kpi-g .kpi-val{color:var(--G)}.kpi-o .kpi-val{color:#B45309}.kpi-p .kpi-val{color:var(--P)}
.kpi-lbl{font-size:.78rem;font-weight:600;color:var(--mu);margin:3px 0 0;text-transform:uppercase;letter-spacing:.06em}
/* ── Activité récente ── */
.activity-item{display:flex;align-items:center;gap:12px;padding:10px 0;border-bottom:1px solid var(--bor)}
.activity-item:last-child{border-bottom:none}
.act-av{width:36px;height:36px;border-radius:50%;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:13px;flex-shrink:0;background:var(--Bm);color:var(--B)}
.act-name{font-weight:600;font-size:14px;color:var(--txt)}
.act-quiz{font-size:.78rem;color:var(--mu)}
.act-score{margin-left:auto;text-align:right}
.act-pct{font-size:1rem;font-weight:800}
.act-time{font-size:.72rem;color:var(--mu)}
/* ── Agent card admin ── */
.agcard{background:var(--sur);border:1.5px solid var(--bor);border-radius:var(--r);padding:12px 14px;margin:6px 0;box-shadow:var(--sh)}
.agtop{display:flex;align-items:center;gap:10px;margin-bottom:8px}
.av{width:36px;height:36px;border-radius:50%;background:var(--Bm);color:var(--B);display:flex;align-items:center;justify-content:center;font-weight:800;font-size:13px;flex-shrink:0}
/* ── Badges ── */
.badge{display:inline-block;padding:2px 9px;border-radius:99px;font-size:.72rem;font-weight:700}
.bg{background:#D1FAE5;color:#065F46}.bo{background:#FEF3C7;color:#92400E}
.bb{background:var(--Bm);color:var(--Bd)}.bgr{background:#F1F5F9;color:#475569}
/* ── Barre de score quiz ── */
.qz-bar-wrap{margin:6px 0}
.qz-bar-row{display:flex;align-items:center;gap:10px;margin:5px 0}
.qz-bar-lbl{font-size:.82rem;font-weight:600;color:var(--txt);min-width:120px;max-width:160px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.qz-bar-bg{flex:1;background:var(--Bm);border-radius:99px;height:10px;overflow:hidden}
.qz-bar-fill{height:10px;border-radius:99px;background:linear-gradient(90deg,var(--B),#60A5FA)}
.qz-bar-val{font-size:.82rem;font-weight:700;color:var(--B);min-width:38px;text-align:right}
/* ── Tabs ── */
.stTabs [data-baseweb="tab-list"]{gap:4px!important;border-bottom:2px solid var(--bor)!important}
.stTabs [data-baseweb="tab"]{font-family:'Outfit',sans-serif!important;font-size:13.5px!important;font-weight:600!important;padding:9px 14px!important;border-radius:8px 8px 0 0!important}
/* ── Expander ── */
div[data-testid="stExpander"]{border:1.5px solid var(--bor)!important;border-radius:var(--r)!important;margin:5px 0!important;overflow:hidden!important;box-shadow:var(--sh)!important}
div[data-testid="stExpander"] summary{font-family:'Outfit',sans-serif!important;font-weight:600!important;font-size:14px!important;padding:13px 16px!important}
div[data-testid="stForm"]{border:none!important;padding:0!important}
hr{border:none!important;border-top:1.5px solid var(--bor)!important;margin:18px 0!important}
.stAlert{border-radius:var(--r)!important;font-family:'Outfit',sans-serif!important}
/* ── Éditeur riche ── */
.rte-toolbar{display:flex;flex-wrap:wrap;gap:6px;margin-bottom:8px;padding:8px 10px;background:var(--bg);border:1.5px solid var(--bor);border-radius:10px 10px 0 0;border-bottom:none}
.rte-toolbar button{background:var(--sur);border:1.5px solid var(--bor);border-radius:6px;padding:4px 10px;font-size:13px;cursor:pointer;font-family:'Outfit',sans-serif;font-weight:600;transition:all .12s;color:var(--txt)}
.rte-toolbar button:hover{background:var(--Bl);border-color:var(--B);color:var(--B)}
.rte-editor{width:100%;min-height:110px;border:1.5px solid var(--bor);border-radius:0 0 10px 10px;padding:10px 12px;font-size:15px;font-family:'Outfit',sans-serif;background:var(--sur);color:var(--txt);outline:none;resize:vertical}
.rte-editor:focus{border-color:var(--B);box-shadow:0 0 0 3px rgba(29,78,216,.1)}
.rte-preview{font-size:15px;font-weight:600;color:var(--txt);line-height:1.5;margin:4px 0 8px;padding:8px 10px;background:var(--Bl);border-radius:8px;border-left:3px solid var(--B)}
/* ── Stats questions ── */
.qstat-card{background:var(--sur);border:1.5px solid var(--bor);border-radius:var(--r);padding:14px 16px;margin:6px 0;box-shadow:var(--sh)}
.qstat-header{display:flex;align-items:flex-start;gap:10px;margin-bottom:8px}
.qstat-num{background:var(--R);color:#fff;font-size:.72rem;font-weight:800;padding:2px 8px;border-radius:99px;flex-shrink:0;margin-top:2px}
.qstat-num-ok{background:var(--G)}
.qstat-num-mid{background:var(--O)}
.qstat-txt{font-size:13.5px;font-weight:600;color:var(--txt);line-height:1.4;flex:1}
.qstat-bar-wrap{display:flex;align-items:center;gap:10px;margin-top:6px}
.qstat-bar-bg{flex:1;background:#FEE2E2;border-radius:99px;height:8px}
.qstat-bar-fill{height:8px;border-radius:99px}
.qstat-meta{display:flex;gap:12px;margin-top:6px;flex-wrap:wrap}
.qstat-pill{font-size:.75rem;font-weight:700;padding:2px 9px;border-radius:99px}
.pill-err{background:#FEE2E2;color:#991B1B}
.pill-ok{background:#D1FAE5;color:#065F46}
.pill-tot{background:var(--Bm);color:var(--Bd)}
</style>""", unsafe_allow_html=True)

init_db()

try:
    from streamlit_autorefresh import st_autorefresh
    _AR = True
except ImportError:
    _AR = False

_D = {
    "page":"home","current_agent":None,"current_quiz":None,
    "quiz_questions":None,"quiz_start_time":None,"quiz_answers":{},
    "session_id":None,"admin_logged":False,"quiz_submitted":False,
    "final_score":None,"edit_quiz_id":None,"adding_q_type":"single",
}
for k,v in _D.items():
    if k not in st.session_state: st.session_state[k]=v


def go(p): st.session_state.page=p; st.rerun()


# ── Persistance de session via URL params (survit au refresh) ─────────────────

def _admin_token():
    """Token de session admin basé sur le mot de passe + date du jour."""
    day = datetime.now().strftime("%Y%m%d")
    raw = f"{config.ADMIN_PASSWORD}{day}"
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def _save_state():
    """Écrit l'état minimal dans les query params."""
    try:
        params = {}
        if st.session_state.current_agent:
            params["a"] = str(st.session_state.current_agent["id"])
        if st.session_state.admin_logged:
            params["t"] = _admin_token()
        if st.session_state.session_id and not st.session_state.quiz_submitted:
            params["s"] = str(st.session_state.session_id)
        # Nettoyer les params obsolètes
        if not st.session_state.current_agent and "a" in st.query_params:
            del st.query_params["a"]
        if not st.session_state.admin_logged and "t" in st.query_params:
            del st.query_params["t"]
        if (not st.session_state.session_id or st.session_state.quiz_submitted) and "s" in st.query_params:
            del st.query_params["s"]
        st.query_params.update(params)
    except Exception:
        pass


def _restore_state():
    """Restaure l'état depuis les query params au démarrage."""
    try:
        qp = st.query_params

        # Restaurer l'agent
        if "a" in qp and not st.session_state.current_agent:
            agent = get_agent_by_id(int(qp["a"]))
            if agent:
                st.session_state.current_agent = agent

        # Restaurer le login admin
        if "t" in qp and not st.session_state.admin_logged:
            if qp["t"] == _admin_token():
                st.session_state.admin_logged = True

        # Restaurer une session de quiz en cours
        if "s" in qp and st.session_state.current_agent and not st.session_state.session_id:
            from database import get_incomplete_session, get_quiz
            agent = st.session_state.current_agent
            # Chercher la session incomplète par ID
            sid = int(qp["s"])
            sess = _fetchone_session(sid)
            if sess and not sess.get("completed"):
                quiz = get_quiz(sess["quiz_id"])
                if quiz:
                    questions = get_questions(quiz["id"])
                    try:
                        raw = sess.get("answers_json") or "{}"
                        answers = json.loads(raw)
                        answers = {int(k) if str(k).isdigit() else k: v for k,v in answers.items()}
                    except Exception:
                        answers = {}
                    remaining = quiz["duree_minutes"]*60 - (time.time() - float(sess.get("start_time_epoch") or time.time()))
                    if remaining > 10:
                        st.session_state.current_quiz = quiz
                        st.session_state.quiz_questions = questions
                        st.session_state.quiz_start_time = float(sess.get("start_time_epoch") or time.time())
                        st.session_state.quiz_answers = answers
                        st.session_state.session_id = sid
                        st.session_state.page = "agent_quiz"
    except Exception:
        pass


def _fetchone_session(sid):
    """Récupère une session par son ID."""
    from database import _fetchone as _db_fetchone
    return _db_fetchone("SELECT * FROM sessions WHERE id=?", (sid,))


# Restaurer au démarrage
_restore_state()


def _generate_quiz_pdf(quiz, questions):
    """Génère un PDF du quiz avec les bonnes réponses via HTML/CSS."""
    from io import BytesIO
    import textwrap

    ICONS = {"single": "⭕", "multiple": "☑️", "numeric": "🔢", "text": "📝"}
    TYPE_LABELS = {"single": "Choix unique", "multiple": "Choix multiple",
                   "numeric": "Valeur numérique", "text": "Texte libre"}

    rows = ""
    for i, q in enumerate(questions):
        pts = f"{q['points']:.0f} pt" + ("s" if q["points"] != 1 else "")
        rows += f"""
        <div class="question">
          <div class="q-header">
            <span class="q-num">Q{i+1}</span>
            <span class="q-type">{ICONS[q["type"]]} {TYPE_LABELS[q["type"]]}</span>
            <span class="q-pts">{pts}</span>
          </div>
          <div class="q-text">{q["texte"]}</div>"""

        if q["type"] in ("single", "multiple"):
            rows += '<div class="options">'
            for o in q["options"]:
                cls = "opt-correct" if o["is_correct"] else "opt-normal"
                icon = "✅" if o["is_correct"] else "☐"
                rows += f'<div class="{cls}">{icon} {o["texte"]}</div>'
            rows += '</div>'
        elif q["type"] == "numeric":
            rows += f'<div class="answer-box">Réponse : <b>{q["reponse_correcte_num"]}</b></div>'
        elif q["type"] == "text":
            raw = q.get("reponse_correcte_txt") or ""
            if raw:
                rows += f'<div class="answer-box">Réponse(s) acceptée(s) : <b>{raw}</b></div>'
            else:
                rows += '<div class="answer-box" style="color:#888">Question ouverte</div>'

        rows += "</div>"

    now = datetime.now().strftime("%d/%m/%Y à %H:%M")
    total_pts = sum(q["points"] for q in questions)
    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;600;700;800&display=swap');
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Outfit', Arial, sans-serif; color: #0F172A; background: #fff; padding: 32px; font-size: 13px; }}
  .header {{ background: linear-gradient(135deg, #1E3A8A, #1D4ED8); color: white; padding: 24px 28px; border-radius: 12px; margin-bottom: 24px; }}
  .header h1 {{ font-size: 22px; font-weight: 800; margin-bottom: 4px; }}
  .header .meta {{ font-size: 12px; opacity: .8; display: flex; gap: 20px; flex-wrap: wrap; margin-top: 8px; }}
  .header .code {{ background: rgba(255,255,255,.2); padding: 3px 12px; border-radius: 99px; font-weight: 700; }}
  .question {{ background: #F8FAFF; border: 1.5px solid #E2E8F4; border-left: 4px solid #1D4ED8; border-radius: 0 10px 10px 0; padding: 14px 16px; margin-bottom: 14px; page-break-inside: avoid; }}
  .q-header {{ display: flex; align-items: center; gap: 8px; margin-bottom: 8px; flex-wrap: wrap; }}
  .q-num {{ background: #EFF6FF; color: #1D4ED8; font-size: 11px; font-weight: 800; padding: 2px 8px; border-radius: 99px; }}
  .q-type {{ color: #64748B; font-size: 11px; }}
  .q-pts {{ background: #F1F5F9; color: #64748B; font-size: 11px; font-weight: 700; padding: 2px 8px; border-radius: 99px; margin-left: auto; }}
  .q-text {{ font-size: 14px; font-weight: 600; margin-bottom: 10px; line-height: 1.45; }}
  .options {{ display: flex; flex-direction: column; gap: 5px; }}
  .opt-normal {{ padding: 7px 12px; background: white; border: 1px solid #E2E8F4; border-radius: 7px; font-size: 13px; }}
  .opt-correct {{ padding: 7px 12px; background: #D1FAE5; border: 1.5px solid #059669; border-radius: 7px; font-size: 13px; font-weight: 600; color: #065F46; }}
  .answer-box {{ background: #D1FAE5; border: 1.5px solid #059669; border-radius: 7px; padding: 7px 12px; color: #065F46; font-size: 13px; }}
  .footer {{ text-align: center; color: #94A3B8; font-size: 11px; margin-top: 24px; padding-top: 16px; border-top: 1px solid #E2E8F4; }}
</style>
</head>
<body>
  <div class="header">
    <h1>{quiz["titre"]}</h1>
    <div class="meta">
      <span class="code">Code : {quiz["code"]}</span>
      <span>⏱ {quiz["duree_minutes"]} minutes</span>
      <span>📝 {len(questions)} question(s)</span>
      <span>🏆 {total_pts:.0f} point(s) au total</span>
    </div>
    {f'<div style="margin-top:8px;font-size:12px;opacity:.75">{quiz["description"]}</div>' if quiz.get("description") else ""}
  </div>
  {rows}
  <div class="footer">Document généré le {now} — QuizAgent CI</div>
</body>
</html>"""
    return html.encode("utf-8")


def topbar(t,i="📝"): st.markdown(f'<div class="topbar"><span style="font-size:1.2rem">{i}</span><span class="tt">{t}</span></div>',unsafe_allow_html=True)
def slbl(t): st.markdown(f'<p class="slbl">{t}</p>',unsafe_allow_html=True)
def badge(t,c="b"): return f'<span class="badge b{c}">{t}</span>'

def rich_text_editor(label, key, default_value="", height=120):
    """
    Éditeur riche 100% Streamlit natif.
    Boutons → modifient session_state directement → st.rerun() → textarea mis à jour.
    Aucun JS, aucun iframe, aucun postMessage.
    """
    if key not in st.session_state:
        st.session_state[key] = default_value

    st.markdown(
        f'<div style="font-size:11px;font-weight:700;color:var(--mu);'
        f'text-transform:uppercase;letter-spacing:.08em;margin-bottom:6px">{label}</div>',
        unsafe_allow_html=True
    )

    # ── Aperçu HTML si le texte contient des balises ──
    current = st.session_state[key]
    if current and ("<b>" in current or "<span" in current or "<br" in current):
        st.markdown(
            f'<div class="rte-preview">Aperçu : {current}</div>',
            unsafe_allow_html=True
        )

    # ── Textarea principal ──
    val = st.text_area(
        " ", value=current, key=f"__ta_{key}", height=height,
        label_visibility="collapsed",
        placeholder="Tapez votre question ici…\nSélectionnez du texte dans ce champ, puis cliquez sur un bouton de mise en forme ci-dessous.",
    )
    # Synchroniser session_state à chaque frappe
    st.session_state[key] = val

    # ── Toolbar : boutons Streamlit natifs ──
    st.markdown('<div style="margin:-8px 0 4px;font-size:11px;color:#64748B">✏️ Sélectionnez du texte ci-dessus, puis :</div>', unsafe_allow_html=True)
    c1,c2,c3,c4,c5,c6 = st.columns(6)

    def _apply(tag_open, tag_close=""):
        """Insère la balise autour du texte sélectionné ou à la fin."""
        txt = st.session_state[key]
        # Sans sélection on ajoute à la fin
        st.session_state[key] = txt + tag_open + ("texte" if tag_close else "") + tag_close

    with c1:
        if st.button("**G**", key=f"_rte_b_{key}", help="Gras", use_container_width=True):
            _apply("<b>", "</b>"); st.rerun()
    with c2:
        if st.button("🔴", key=f"_rte_r_{key}", help="Rouge", use_container_width=True):
            _apply('<span style="color:#DC2626">', "</span>"); st.rerun()
    with c3:
        if st.button("🔵", key=f"_rte_bl_{key}", help="Bleu", use_container_width=True):
            _apply('<span style="color:#1D4ED8">', "</span>"); st.rerun()
    with c4:
        if st.button("🟢", key=f"_rte_g_{key}", help="Vert", use_container_width=True):
            _apply('<span style="color:#059669">', "</span>"); st.rerun()
    with c5:
        if st.button("🟠", key=f"_rte_o_{key}", help="Orange", use_container_width=True):
            _apply('<span style="color:#D97706">', "</span>"); st.rerun()
    with c6:
        if st.button("↵", key=f"_rte_br_{key}", help="Saut de ligne", use_container_width=True):
            _apply("<br>"); st.rerun()

    return st.session_state[key]

def kpi_grid(items):
    """items = [(icon, value, label, color_class), ...]"""
    cols = "".join(
        f'<div class="kpi kpi-{c}"><div class="kpi-icon">{i}</div>'
        f'<p class="kpi-val">{v}</p><p class="kpi-lbl">{l}</p></div>'
        for i,v,l,c in items
    )
    st.markdown(f'<div class="kpi-grid">{cols}</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  HOME
# ══════════════════════════════════════════════════════════════

def render_home():
    org = f"<br><small style='opacity:.7'>{config.ORG_NAME}</small>" if config.ORG_NAME else ""
    st.markdown(
        f'<div class="hero"><div class="hi">📝</div>'
        f'<h1 class="ht">{config.APP_TITLE}</h1>'
        f'<p class="hs">{config.APP_SUBTITLE}{org}</p></div>',
        unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        if st.button("👤  Je suis Agent",key="home_agent",use_container_width=True,type="primary"): go("agent_search")
    with c2:
        if st.button("⚙️  Administration",key="home_admin",use_container_width=True): go("admin_login")


# ══════════════════════════════════════════════════════════════
#  AGENT — Recherche
# ══════════════════════════════════════════════════════════════

def render_agent_search():
    topbar("Identification","👤")
    slbl("Tapez votre nom ou matricule")
    q = st.text_input(" ", placeholder="Ex : Konan, Diallo, AGT001…",
                       label_visibility="collapsed", key="sq")
    if st.button("← Retour",key="as_back",use_container_width=True): go("home")
    q = (q or "").strip()
    if len(q)<2:
        if len(q)==1: st.caption("Continuez à taper…")
        return
    agents = search_agents(q)
    if not agents:
        st.warning("Aucun agent trouvé. Vérifiez l'orthographe ou contactez votre superviseur.")
        return
    slbl(f"{len(agents)} résultat(s) — appuyez sur votre nom")
    for a in agents:
        lbl = f"👤  {a['nom']} {a['prenom']}".strip()
        if a["matricule"]: lbl += f"  ·  {a['matricule']}"
        if st.button(lbl,key=f"sel_{a['id']}",use_container_width=True,type="primary"):
            st.session_state.current_agent=a; _save_state(); go("agent_quiz_code")


# ══════════════════════════════════════════════════════════════
#  AGENT — Code quiz
# ══════════════════════════════════════════════════════════════

def render_agent_quiz_code():
    agent=st.session_state.current_agent
    if not agent: go("agent_search"); return
    topbar("Accès au Quiz","📋")
    init=(agent["nom"][0]+(agent["prenom"][0] if agent["prenom"] else "")).upper()
    st.markdown(
        f'<div class="card" style="display:flex;align-items:center;gap:12px;margin-bottom:18px">'
        f'<div class="av" style="width:44px;height:44px;font-size:15px">{init}</div>'
        f'<div><div style="font-size:1.05rem;font-weight:700">{agent["nom"]} {agent["prenom"]}</div>'
        f'<div style="color:var(--mu);font-size:.83rem">{agent["matricule"] or "Agent"}</div></div></div>',
        unsafe_allow_html=True)
    slbl("Code du quiz communiqué par votre superviseur")
    code=st.text_input(" ",placeholder="Ex : QZ001",max_chars=20,label_visibility="collapsed")
    c1,c2=st.columns([1,2])
    with c1:
        if st.button("← Retour",key="aqc_back",use_container_width=True): go("agent_search")
    with c2:
        if st.button("▶  Commencer",key="aqc_start",type="primary",use_container_width=True):
            _start_quiz(agent,code)


def _start_quiz(agent, code):
    if not (code or "").strip(): st.warning("Saisissez un code."); return
    quiz=get_quiz_by_code(code.strip())
    if not quiz: st.error("Code incorrect ou quiz désactivé."); return

    # ── Vérifier session en cours (pour restore après refresh) ──
    incomplete=get_incomplete_session(agent["id"],quiz["id"])
    if incomplete and incomplete.get("start_time_epoch"):
        epoch=incomplete["start_time_epoch"]
        remaining=quiz["duree_minutes"]*60-(time.time()-epoch)
        if remaining>15:  # encore du temps
            try:
                raw=incomplete.get("answers_json") or "{}"
                answers=json.loads(raw)
                answers={int(k) if str(k).isdigit() else k: v for k,v in answers.items()}
            except Exception:
                answers={}
            questions=get_questions(quiz["id"])  # ordre non aléatoire à la restauration
            st.session_state.current_quiz=quiz
            st.session_state.quiz_questions=questions
            st.session_state.quiz_start_time=epoch
            st.session_state.quiz_answers=answers
            st.session_state.quiz_submitted=False
            st.session_state.session_id=incomplete["id"]
            st.toast("🔄 Quiz en cours restauré — continuez !", icon="✅")
            go("agent_quiz"); return
        else:
            # temps écoulé sur session incomplète → auto-soumettre
            try:
                raw=incomplete.get("answers_json") or "{}"
                answers=json.loads(raw)
                answers={int(k) if str(k).isdigit() else k: v for k,v in answers.items()}
                questions=get_questions(quiz["id"])
                score,records=_calc(questions,answers)
                submit_session(incomplete["id"],score,records)
            except Exception: pass
            st.warning("Votre temps était écoulé. Votre session a été automatiquement soumise.")
            return

    if session_already_completed(agent["id"],quiz["id"]):
        st.warning("Vous avez déjà soumis ce quiz. Contactez votre superviseur."); return

    shuffled=bool(quiz.get("randomize_questions",0))
    questions=get_questions(quiz["id"],shuffled=shuffled)
    if not questions: st.error("Ce quiz ne contient pas encore de questions."); return

    max_score=sum(q["points"] for q in questions)
    epoch=time.time()
    sid=create_session(agent["id"],quiz["id"],max_score,epoch)
    st.session_state.current_quiz=quiz
    st.session_state.quiz_questions=questions
    st.session_state.quiz_start_time=epoch
    st.session_state.quiz_answers={}
    st.session_state.quiz_submitted=False
    st.session_state.session_id=sid
    go("agent_quiz")


# ══════════════════════════════════════════════════════════════
#  AGENT — Quiz (timer fixe)
# ══════════════════════════════════════════════════════════════

def _calc(questions, answers, malus_actif=False, malus_points=0.0):
    score,records=0.0,[]
    for q in questions:
        qid=q["id"]; resp=answers.get(qid,""); ok=0
        if q["type"]=="single":
            cor=[o["id"] for o in q["options"] if o["is_correct"]]
            if resp in cor: ok=1
            elif resp not in ("","",None) and malus_actif: ok=-1
        elif q["type"]=="multiple":
            cor=set(o["id"] for o in q["options"] if o["is_correct"])
            giv=set(resp) if isinstance(resp,list) else set()
            if giv==cor and cor: ok=1
            elif giv and malus_actif: ok=-1
        elif q["type"]=="numeric":
            try:
                exp=float(q["reponse_correcte_num"]) if q["reponse_correcte_num"] is not None else None
                giv=float(str(resp).replace(",",".")) if str(resp).strip() else None
                if exp is not None and giv is not None and abs(giv-exp)<0.01: ok=1
                elif giv is not None and malus_actif: ok=-1
            except: ok=0
        elif q["type"]=="text":
            raw=(q.get("reponse_correcte_txt") or "").strip()
            if raw:
                if str(resp).lower().strip() in [a.lower().strip() for a in raw.split("|") if a.strip()]: ok=1
                elif str(resp).strip() and malus_actif: ok=-1
            else: ok=1
        if ok==1:  score+=q["points"]
        if ok==-1: score-=(malus_points if malus_points>0 else q["points"])
        records.append({"question_id":qid,
                         "reponse":json.dumps(resp,ensure_ascii=False) if isinstance(resp,list) else str(resp),
                         "is_correct":1 if ok==1 else 0})
    return max(0.0, score), records


def _submit():
    if st.session_state.quiz_submitted: return
    ql=st.session_state.quiz_questions
    qz=st.session_state.current_quiz
    malus_actif=bool(qz.get("malus_actif",0))
    malus_pts=float(qz.get("malus_points",0) or 0)
    score,records=_calc(ql,st.session_state.quiz_answers,malus_actif,malus_pts)
    submit_session(st.session_state.session_id,score,records)
    st.session_state.final_score={"score":score,"max_score":sum(q["points"] for q in ql)}
    st.session_state.quiz_submitted=True
    st.session_state.result_session_id=st.session_state.session_id


def render_agent_quiz():
    if st.session_state.quiz_submitted: go("agent_result"); return

    quiz=st.session_state.current_quiz
    questions=st.session_state.quiz_questions
    agent=st.session_state.current_agent

    if _AR: st_autorefresh(interval=2000,key="qr")

    remaining=max(0,quiz["duree_minutes"]*60-(time.time()-st.session_state.quiz_start_time))

    if remaining<=0:
        _submit(); st.session_state.page="agent_result"; st.rerun(); return

    mins,secs=int(remaining//60),int(remaining%60)
    if remaining>300:  tcls,tlbl,tico="tn","Temps restant","⏱"
    elif remaining>60: tcls,tlbl,tico="tw","Moins de 5 min !","⚠️"
    else:              tcls,tlbl,tico="td","Dépêchez-vous !","🚨"

    # ── TIMER FIXE — position:fixed via JS pour garantir le fonctionnement ──
    st.markdown(f"""
    <div id="quiz-timer-bar">
      <div class="inner">
        <div class="tbox {tcls}">
          <span class="ticon">{tico}</span>
          <div><p class="ttime">{mins:02d}:{secs:02d}</p><p class="tlbl">{tlbl}</p></div>
        </div>
      </div>
    </div>
    <div style="height:90px"></div>
    <script>
    (function(){{
      var el=document.getElementById('quiz-timer-bar');
      if(el && el.parentNode!==document.body){{
        document.body.insertBefore(el,document.body.firstChild);
      }}
    }})();
    </script>
    """, unsafe_allow_html=True)

    answered=sum(1 for v in st.session_state.quiz_answers.values() if v not in("","[],",None,[]))
    pct=answered/len(questions) if questions else 0
    st.markdown(
        f'<div class="prog">'
        f'<div class="prog-row"><span><b>{quiz["titre"]}</b></span><span>{answered}/{len(questions)} répondue(s)</span></div>'
        f'<div class="prog-bg"><div class="prog-fill" style="width:{pct*100:.0f}%"></div></div>'
        f'</div>', unsafe_allow_html=True)

    ICONS={"single":"⭕","multiple":"☑️","numeric":"🔢","text":"📝"}
    HINTS={"single":"Une seule réponse","multiple":"Plusieurs réponses","numeric":"Entrez un nombre","text":"Réponse libre"}

    for i,q in enumerate(questions):
        qid=q["id"]
        pts=f"{q['points']:.0f} pt"+("s" if q["points"]!=1 else "")
        st.markdown(
            f'<div class="qcard">'
            f'<div class="qmeta"><span class="qnum">Q{i+1}</span>'
            f'<span class="qtype">{ICONS[q["type"]]} {HINTS[q["type"]]}</span>'
            f'<span class="qpts">{pts}</span></div>'
            f'<p class="qtxt">{_fmt(q["texte"])}</p></div>', unsafe_allow_html=True)

        if q["type"]=="single":
            opts=q["options"]; cur=st.session_state.quiz_answers.get(qid)
            idx=next((j for j,o in enumerate(opts) if o["id"]==cur),None)
            sel=st.radio(" ",range(len(opts)),format_func=lambda x,_o=opts:_o[x]["texte"],
                         index=idx,key=f"q_{qid}",label_visibility="collapsed")
            if sel is not None: st.session_state.quiz_answers[qid]=opts[sel]["id"]

        elif q["type"]=="multiple":
            cur=st.session_state.quiz_answers.get(qid,[])
            sel=[]
            for o in q["options"]:
                if st.checkbox(o["texte"],value=(o["id"] in cur),key=f"q_{qid}_{o['id']}"): sel.append(o["id"])
            st.session_state.quiz_answers[qid]=sel

        elif q["type"]=="numeric":
            val=st.session_state.quiz_answers.get(qid,"")
            v=st.text_input(" ",value=str(val) if val!="" else "",key=f"q_{qid}",
                             placeholder="Entrez un nombre…",label_visibility="collapsed")
            st.session_state.quiz_answers[qid]=v

        elif q["type"]=="text":
            val=st.session_state.quiz_answers.get(qid,"")
            v=st.text_area(" ",value=val,key=f"q_{qid}",height=80,
                            placeholder="Votre réponse…",label_visibility="collapsed")
            st.session_state.quiz_answers[qid]=v

        st.markdown("")

    st.markdown("---")
    if st.button("✅  Soumettre mes réponses",key="quiz_submit",type="primary",use_container_width=True):
        _submit()
        go("agent_result")
        return

    # ── Sauvegarde auto des réponses (uniquement si pas en train de soumettre) ──
    if st.session_state.session_id and not st.session_state.quiz_submitted:
        try:
            save_quiz_progress(
                st.session_state.session_id,
                json.dumps(st.session_state.quiz_answers),
                st.session_state.quiz_start_time
            )
        except Exception: pass


# ══════════════════════════════════════════════════════════════
#  AGENT — Résultat
# ══════════════════════════════════════════════════════════════

def render_agent_result():
    res=st.session_state.final_score; quiz=st.session_state.current_quiz; agent=st.session_state.current_agent
    if not res or not agent or not quiz: go("home"); return
    show_score=bool(quiz.get("show_score",1))
    score,max_sc=res["score"],res["max_score"]
    pct=(score/max_sc*100) if max_sc>0 else 0
    if show_score:
        emoji="🏆" if pct>=80 else("👍" if pct>=60 else "📚")
        st.markdown(
            f'<div class="scard"><div class="se">{emoji}</div>'
            f'<p class="sn">{agent["nom"]} {agent["prenom"]}</p>'
            f'<p class="sv">{score:.1f} / {max_sc:.1f}</p>'
            f'<p class="sp">{pct:.0f} %</p><p class="sq">{quiz["titre"]}</p></div>',
            unsafe_allow_html=True)
        if pct>=80: st.success("Excellent résultat ! Félicitations ! 🎉")
        elif pct>=60: st.info("Bon résultat. Continuez vos efforts !")
        else: st.warning("Il faut encore travailler ce sujet. Courage !")
    else:
        st.markdown(
            f'<div class="subcard"><div style="font-size:2.5rem;margin-bottom:8px">✅</div>'
            f'<div style="font-size:1.1rem;font-weight:700;color:var(--Bd)">Réponses enregistrées</div>'
            f'<div style="color:var(--mu);font-size:.9rem;margin-top:6px">{agent["nom"]} {agent["prenom"]} · {quiz["titre"]}</div></div>',
            unsafe_allow_html=True)
    st.markdown(f'<p style="text-align:center;color:var(--mu);font-size:.82rem;margin:6px 0 14px">Soumis le {datetime.now().strftime("%d/%m/%Y à %H:%M")}</p>',unsafe_allow_html=True)
    if quiz.get("show_correction", 0):
        if st.button("📖  Voir mon corrigé",key="res_corr",type="primary",use_container_width=True):
            go("agent_correction")
    if st.button("🏠  Retour à l'accueil",key="result_home",use_container_width=True):
        for k in("current_quiz","quiz_questions","quiz_start_time","quiz_answers","session_id","quiz_submitted","final_score"): st.session_state[k]=_D[k]
        st.session_state.current_agent=None; _save_state(); go("home")


# ══════════════════════════════════════════════════════════════
#  ADMIN — Connexion
# ══════════════════════════════════════════════════════════════

def render_admin_login():
    if st.session_state.admin_logged: go("admin_dashboard"); return
    topbar("Administration","⚙️")
    st.markdown('<div class="card">',unsafe_allow_html=True)
    st.markdown("#### 🔐 Connexion administrateur")
    pwd=st.text_input(" ",type="password",placeholder="Mot de passe…",label_visibility="collapsed")
    c1,c2=st.columns([1,2])
    with c1:
        if st.button("← Retour",key="al_back",use_container_width=True): go("home")
    with c2:
        if st.button("Se connecter",key="al_login",type="primary",use_container_width=True):
            if pwd==config.ADMIN_PASSWORD: st.session_state.admin_logged=True; _save_state(); go("admin_dashboard")
            else: st.error("Mot de passe incorrect.")
    st.markdown('</div>',unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════
#  ADMIN — Dashboard
# ══════════════════════════════════════════════════════════════

def render_admin_dashboard():
    if not st.session_state.admin_logged: go("admin_login"); return
    topbar(config.APP_TITLE+"  ·  Tableau de bord","⚙️")

    t0,t1,t2,t3,t4,t5=st.tabs(["📊 Vue d'ensemble","👥 Agents","📝 Quiz","📋 Résultats","📉 Stats questions","🔎 Surveillance"])
    with t0: _tab_overview()
    with t1: _tab_agents()
    with t2: _tab_quizzes()
    with t3: _tab_results()
    with t4: _tab_question_stats()
    with t5: _tab_surveillance()

    st.markdown("---")
    if st.button("🚪  Déconnexion",key="admin_logout",use_container_width=True): st.session_state.admin_logged=False; _save_state(); go("home")


# ── Vue d'ensemble ────────────────────────────────────────────

def _tab_overview():
    stats=get_stats()

    kpi_grid([
        ("👥", stats["total_agents"],   "Agents total",  "b"),
        ("✅", stats["pub_agents"],      "Publiés",       "g"),
        ("📝", stats["active_quizzes"],  "Quiz actifs",   "p"),
        ("📊", stats["total_submissions"],"Soumissions",  "o"),
    ])

    # Score moyen global
    avg=stats["avg_score"]
    c1,c2=st.columns(2)
    with c1:
        st.markdown(
            f'<div class="card" style="text-align:center">'
            f'<div style="font-size:.72rem;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px">Score moyen global</div>'
            f'<div style="font-size:2.4rem;font-weight:900;color:{"var(--G)" if avg>=70 else "var(--O)" if avg>=50 else "var(--R)"}">{avg:.1f}%</div>'
            f'</div>',unsafe_allow_html=True)
    with c2:
        # Taux de complétion : sessions complétées / agents publiés
        rate=min(100,round(stats["total_submissions"]/max(1,stats["pub_agents"])*100))
        st.markdown(
            f'<div class="card" style="text-align:center">'
            f'<div style="font-size:.72rem;font-weight:700;color:var(--mu);text-transform:uppercase;letter-spacing:.08em;margin-bottom:8px">Agents ayant composé</div>'
            f'<div style="font-size:2.4rem;font-weight:900;color:var(--B)">{rate}%</div>'
            f'</div>',unsafe_allow_html=True)

    # Scores par quiz
    if stats["per_quiz"]:
        st.markdown('<p class="slbl">Scores moyens par quiz</p>',unsafe_allow_html=True)
        bars=""
        for q in stats["per_quiz"]:
            pct=min(100,q["avg_pct"] or 0)
            col="var(--G)" if pct>=70 else("var(--O)" if pct>=50 else "var(--R)")
            bars+=f'<div class="qz-bar-row"><span class="qz-bar-lbl" title="{q["titre"]}">{q["titre"][:20]}</span><div class="qz-bar-bg"><div class="qz-bar-fill" style="width:{pct:.0f}%;background:linear-gradient(90deg,{col},{col}88)"></div></div><span class="qz-bar-val">{pct:.0f}%</span></div>'
        st.markdown(f'<div class="qz-bar-wrap">{bars}</div>',unsafe_allow_html=True)

    # Activité récente
    if stats["recent"]:
        st.markdown('<p class="slbl">10 dernières soumissions</p>',unsafe_allow_html=True)
        items=""
        for r in stats["recent"]:
            pct_r=(r["score"]/r["max_score"]*100) if r["max_score"]>0 else 0
            col_r="var(--G)" if pct_r>=70 else("var(--O)" if pct_r>=50 else "var(--R)")
            init_r=(r["nom"][0]+(r["prenom"][0] if r["prenom"] else "")).upper()
            # Format time
            try:
                dt=datetime.fromisoformat(r["completed_at"])
                diff=datetime.now()-dt
                if diff.seconds<3600:    tstr=f"Il y a {diff.seconds//60} min"
                elif diff.days==0:       tstr=f"Il y a {diff.seconds//3600}h"
                else:                    tstr=dt.strftime("%d/%m %H:%M")
            except: tstr=r["completed_at"][:16] if r["completed_at"] else "—"
            items+=(f'<div class="activity-item">'
                    f'<div class="act-av">{init_r}</div>'
                    f'<div><div class="act-name">{r["nom"]} {r["prenom"]}</div>'
                    f'<div class="act-quiz">{r["titre"]}</div></div>'
                    f'<div class="act-score">'
                    f'<div class="act-pct" style="color:{col_r}">{pct_r:.0f}%</div>'
                    f'<div class="act-time">{tstr}</div></div></div>')
        st.markdown(f'<div class="card">{items}</div>',unsafe_allow_html=True)
    else:
        st.info("Aucune soumission pour l'instant.")

    # Actions rapides
    st.markdown('<p class="slbl">Actions rapides</p>',unsafe_allow_html=True)
    c1,c2=st.columns(2)
    with c1:
        if st.button("➕  Créer un quiz",key="ovw_create_quiz",use_container_width=True,type="primary"):
            st.session_state.edit_quiz_id=None; go("admin_quiz_edit")
    with c2:
        if st.button("✅  Publier tous les agents",key="ovw_pub_all",use_container_width=True):
            publish_all_agents(); st.rerun()


# ── Agents ────────────────────────────────────────────────────

def _ndf(df):
    df.columns=[str(c).lower().strip() for c in df.columns]
    nom=next((c for c in ["nom","name","lastname","noms"] if c in df.columns),None)
    if not nom: return None
    pre=next((c for c in ["prenom","prénom","firstname","prenoms"] if c in df.columns),None)
    mat=next((c for c in ["matricule","id","code","identifiant"] if c in df.columns),None)
    out=pd.DataFrame()
    out["nom"]=df[nom]; out["prenom"]=df[pre] if pre else ""; out["matricule"]=df[mat] if mat else ""
    return out


def _tab_agents():
    s1,s2,s3=st.tabs(["✏️ Ajouter","📥 Importer","📋 Liste"])
    with s1:
        with st.form("fag",clear_on_submit=True):
            c1,c2,c3=st.columns(3)
            with c1: nom=st.text_input("Nom *")
            with c2: pre=st.text_input("Prénom")
            with c3: mat=st.text_input("Matricule")
            if st.form_submit_button("Ajouter",type="primary",use_container_width=True):
                if not nom.strip(): st.error("Nom obligatoire.")
                elif add_agent_manual(nom.strip(),pre.strip(),mat.strip()): st.success(f"{nom} ajouté."); st.rerun()
                else: st.warning("Cet agent existe déjà.")
    with s2:
        st.markdown("**Fichier CSV ou Excel**"); st.caption("Colonnes : `nom`, `prenom`, `matricule`")
        up=st.file_uploader("Fichier",type=["csv","xlsx","xls"],label_visibility="collapsed")
        if up:
            try:
                df_r=pd.read_csv(up) if up.name.lower().endswith(".csv") else pd.read_excel(up)
                df=_ndf(df_r)
                if df is None: st.error("Colonne 'nom' introuvable.")
                else:
                    st.dataframe(df.head(5),use_container_width=True)
                    st.caption(f"{len(df)} agent(s)")
                    if st.button("✅  Importer",key="ag_import_file",type="primary",use_container_width=True):
                        n=upsert_agents(df); st.success(f"{n} importé(s)."); st.rerun()
            except Exception as e: st.error(f"Erreur : {e}")
        st.markdown("---"); st.markdown("**Google Sheets (lien public)**")
        gs=st.text_input("URL",placeholder="https://docs.google.com/spreadsheets/d/…",label_visibility="collapsed")
        if st.button("Importer depuis Google Sheets",key="ag_import_gs",use_container_width=True): _gs(gs)
    with s3:
        ags=get_all_agents()
        if not ags: st.info("Aucun agent."); return
        nb=sum(1 for a in ags if a["published"])
        st.markdown(f'<p style="color:var(--mu);font-size:.85rem;margin-bottom:10px"><b>{len(ags)}</b> agent(s) &nbsp;·&nbsp;{badge(f"{nb} publié(s)","g")} &nbsp;{badge(f"{len(ags)-nb} brouillon(s)","o")}</p>',unsafe_allow_html=True)
        c1,c2=st.columns(2)
        with c1:
            if st.button("✅ Publier tous",key="ag_pub_all",use_container_width=True,type="primary"): publish_all_agents(); st.rerun()
        with c2:
            if st.button("⛔ Dépublier tous",key="ag_unpub_all",use_container_width=True): unpublish_all_agents(); st.rerun()
        st.markdown("---")
        filtre=st.text_input("🔍 Filtrer",placeholder="Nom, matricule…",key="af",label_visibility="collapsed")
        shown=[a for a in ags if not filtre or filtre.lower() in f"{a['nom']} {a['prenom']} {a['matricule']}".lower()]
        for a in shown:
            init=(a["nom"][0]+(a["prenom"][0] if a["prenom"] else "")).upper()
            pub_b=badge("✓ Publié","g") if a["published"] else badge("Brouillon","o")
            mat_s=f'  ·  <span style="color:var(--mu)">{a["matricule"]}</span>' if a["matricule"] else ""
            st.markdown(f'<div class="agcard"><div class="agtop"><div class="av">{init}</div><div><b>{a["nom"]} {a["prenom"]}</b>{mat_s}<br>{pub_b}</div></div>',unsafe_allow_html=True)
            c1,c2=st.columns([3,1])
            with c1:
                lbl="🔓 Dépublier" if a["published"] else "✅ Publier"
                if st.button(lbl,key=f"p_{a['id']}",use_container_width=True): set_agent_published(a["id"],0 if a["published"] else 1); st.rerun()
            with c2:
                if st.button("🗑️",key=f"d_{a['id']}",use_container_width=True): delete_agent(a["id"]); st.rerun()
            st.markdown('</div>',unsafe_allow_html=True)


def _gs(url):
    if not (url or "").strip(): st.warning("Collez une URL."); return
    try:
        m=re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)",url)
        if not m: st.error("URL invalide."); return
        sid=m.group(1); gm=re.search(r"gid=(\d+)",url); gid=gm.group(1) if gm else "0"
        r=requests.get(f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv&gid={gid}",timeout=15); r.raise_for_status()
        df=_ndf(pd.read_csv(io.StringIO(r.text)))
        if df is None: st.error("Colonne 'nom' introuvable."); return
        n=upsert_agents(df); st.success(f"{n} agent(s) importé(s)."); st.rerun()
    except Exception as e: st.error(f"Erreur : {e}")


# ── Quiz ──────────────────────────────────────────────────────

def _tab_quizzes():
    if st.button("➕  Créer un quiz",key="qz_create_quiz",type="primary",use_container_width=True):
        st.session_state.edit_quiz_id=None; go("admin_quiz_edit")
    st.markdown("---")
    quizzes=get_all_quizzes()
    if not quizzes: st.info("Aucun quiz."); return
    for quiz in quizzes:
        nb=len(get_questions(quiz["id"]))
        s=badge("Actif","g") if quiz["actif"] else badge("Inactif","gr")
        sc=badge("Score visible","b") if quiz.get("show_score",1) else badge("Score masqué","o")
        rn=badge("Aléatoire","b") if quiz.get("randomize_questions",0) else ""
        with st.expander(f"{'🟢' if quiz['actif'] else '⚫'} {quiz['titre']}  ·  `{quiz['code']}`"):
            st.markdown(f"{s} &nbsp;{sc} &nbsp;{rn} &nbsp;{badge(str(quiz['duree_minutes'])+' min','gr')} &nbsp;{badge(str(nb)+' question(s)','gr')}",unsafe_allow_html=True)
            if quiz["description"]: st.caption(quiz["description"])
            c1,c2,c3=st.columns(3)
            with c1:
                if st.button("✏️ Modifier",key=f"eq_{quiz['id']}",use_container_width=True): st.session_state.edit_quiz_id=quiz["id"]; go("admin_quiz_edit")
            with c2:
                if st.button("🟢 Activer" if not quiz["actif"] else "🔴 Désactiver",key=f"tq_{quiz['id']}",use_container_width=True):
                    update_quiz(quiz["id"],quiz["titre"],quiz["code"],quiz["duree_minutes"],quiz["description"],0 if quiz["actif"] else 1,quiz.get("show_score",1),quiz.get("randomize_questions",0)); st.rerun()
            with c3:
                if st.button("🗑 Suppr.",key=f"dq_{quiz['id']}",use_container_width=True): delete_quiz(quiz["id"]); st.rerun()
            # Export PDF corrigé
            if nb > 0:
                qz_questions = get_questions(quiz["id"])
                pdf_bytes = _generate_quiz_pdf(quiz, qz_questions)
                st.download_button(
                    "📄  Télécharger le corrigé PDF",
                    data=pdf_bytes,
                    file_name=f"corrige_{quiz['code']}.html",
                    mime="text/html",
                    key=f"pdf_{quiz['id']}",
                    use_container_width=True,
                )


# ── Résultats ─────────────────────────────────────────────────

def _tab_results():
    quizzes = get_all_quizzes()
    opts = {0: "— Tous les quiz —"}
    opts.update({q["id"]: f"{q['titre']} ({q['code']})" for q in quizzes})
    sel = st.selectbox("Quiz", list(opts.keys()), format_func=lambda x: opts[x],
                       label_visibility="collapsed", key="res_sel")
    results = get_results(sel if sel else None)
    if not results:
        st.info("Aucun résultat disponible."); return

    # ── Tableau résumé ──
    df = pd.DataFrame(results)
    df["pct"] = (df["score"] / df["max_score"] * 100).round(1)
    df["agent"] = df["nom"] + " " + df["prenom"]
    disp = df[["agent","matricule","quiz_titre","score","max_score","pct","completed_at"]].copy()
    disp.columns = ["Agent","Matricule","Quiz","Score","Sur","%","Date"]
    st.dataframe(disp, use_container_width=True)

    # ── Export choix ──
    mode = st.radio("Export", ["Résumé uniquement", "Avec réponses par question"],
                    horizontal=True, key="res_export_mode")

    if mode == "Résumé uniquement":
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            disp.to_excel(w, index=False, sheet_name="Résultats")
            ws = w.sheets["Résultats"]
            from openpyxl.styles import PatternFill, Font
            for row in ws.iter_rows(min_row=2):
                pct_val = row[5].value or 0
                if pct_val >= 70: row[5].font = Font(color="059669", bold=True)
                elif pct_val < 40: row[5].font = Font(color="DC2626", bold=True)
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = min(
                    max(len(str(c.value or ""))+4 for c in col), 40)
        buf.seek(0)
        st.download_button("📥 Télécharger Excel (résumé)",
            data=buf.getvalue(),
            file_name=f"resultats_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, key="dl_res_summary")

    else:
        # Export détaillé : une ligne par question par agent
        if st.button("🔄 Générer l'export détaillé", key="gen_detail", type="primary", use_container_width=True):
            with st.spinner("Collecte des réponses en cours…"):
                rows_detail = []
                for r in results:
                    sid = r.get("id")
                    if not sid: continue
                    answers = get_session_answers(sid)
                    pct = round(r["score"]/r["max_score"]*100, 1) if r.get("max_score",0)>0 else 0
                    for a in answers:
                        # Réponse lisible
                        resp_raw = a.get("reponse","")
                        try:
                            import json as _j; rp = _j.loads(resp_raw)
                        except: rp = resp_raw
                        if a["type"] in ("single","multiple"):
                            opts_q = a.get("options",[])
                            rids = rp if isinstance(rp,list) else ([rp] if rp else [])
                            resp_txt = " / ".join(o["texte"] for o in opts_q if o["id"] in rids) or "—"
                            cor_txt  = " / ".join(o["texte"] for o in opts_q if o["is_correct"])
                        elif a["type"] == "numeric":
                            resp_txt = str(resp_raw) if str(resp_raw).strip() else "—"
                            cor_txt  = str(a.get("reponse_correcte_num",""))
                        else:
                            resp_txt = str(resp_raw) if str(resp_raw).strip() else "—"
                            cor_txt  = (a.get("reponse_correcte_txt") or "Question ouverte")

                        rows_detail.append({
                            "Agent":        f"{r['nom']} {r['prenom']}",
                            "Matricule":    r.get("matricule",""),
                            "Quiz":         r.get("quiz_titre",""),
                            "Score total":  r["score"],
                            "Sur":          r["max_score"],
                            "% total":      pct,
                            "Q N°":         a.get("num", ""),
                            "Question":     _strip_html(a["texte"])[:100],
                            "Type":         a["type"],
                            "Pts question": a["points"],
                            "Réponse agent":resp_txt,
                            "Bonne réponse":cor_txt,
                            "Correct":      "✅" if a["is_correct"] else "❌",
                            "Date":         r.get("completed_at",""),
                        })

                if not rows_detail:
                    st.warning("Aucune réponse détaillée disponible."); return

                df_det = pd.DataFrame(rows_detail)
                buf = io.BytesIO()
                with pd.ExcelWriter(buf, engine="openpyxl") as w:
                    # Feuille résumé
                    disp.to_excel(w, index=False, sheet_name="Résumé")
                    # Feuille détail
                    df_det.to_excel(w, index=False, sheet_name="Réponses détaillées")
                    from openpyxl.styles import PatternFill, Font
                    for sheet_name in ["Résumé", "Réponses détaillées"]:
                        ws = w.sheets[sheet_name]
                        for col in ws.columns:
                            ws.column_dimensions[col[0].column_letter].width = min(
                                max(len(str(c.value or ""))+4 for c in col), 50)
                    # Colorier la feuille détail
                    ws2 = w.sheets["Réponses détaillées"]
                    for row in ws2.iter_rows(min_row=2):
                        cor_cell = row[12]  # colonne Correct
                        if cor_cell.value == "✅":
                            for cell in row: cell.fill = PatternFill("solid", fgColor="D1FAE5")
                        elif cor_cell.value == "❌":
                            for cell in row: cell.fill = PatternFill("solid", fgColor="FEE2E2")
                buf.seek(0)
                st.success(f"✅ {len(rows_detail)} lignes générées")
                st.download_button("📥 Télécharger Excel (détaillé)",
                    data=buf.getvalue(),
                    file_name=f"resultats_detail_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="dl_res_detail")


# ── Stats par question ─────────────────────────────────────────

def _strip_html(text):
    """Supprime les balises HTML pour l'affichage en texte brut (export)."""
    import re
    return re.sub(r'<[^>]+>', '', str(text or ""))


def _tab_question_stats():
    quizzes = get_all_quizzes()
    opts = {0: "— Tous les quiz —"}
    opts.update({q["id"]: f"{q['titre']} ({q['code']})" for q in quizzes})
    sel = st.selectbox("Filtrer par quiz", list(opts.keys()), format_func=lambda x: opts[x],
                       label_visibility="collapsed", key="qs_filter")
    stats = get_question_stats(sel if sel else None)
    if not stats:
        st.info("Aucune donnée disponible. Les stats apparaissent après les premières soumissions.")
        return

    total_q = len(stats)
    erreurs_hautes = sum(1 for s in stats if (s["taux_erreur"] or 0) >= 50)

    kpi_grid([
        ("❓", total_q,        "Questions analysées", "b"),
        ("🚨", erreurs_hautes, "Questions difficiles (≥50% d'erreurs)", "o"),
        ("📊", f"{sum(s['total_reponses'] or 0 for s in stats)}", "Réponses totales", "p"),
        ("✅", f"{sum(s['bonnes_reponses'] or 0 for s in stats)}", "Bonnes réponses", "g"),
    ])

    slbl(f"Questions classées par taux d'erreur (du + difficile au + facile)")

    for i, s in enumerate(stats):
        taux = float(s["taux_erreur"] or 0)
        total = int(s["total_reponses"] or 0)
        bonnes = int(s["bonnes_reponses"] or 0)
        mauvaises = int(s["mauvaises_reponses"] or 0)

        if taux >= 50:   badge_cls = "qstat-num";     bar_col = "#DC2626"
        elif taux >= 25: badge_cls = "qstat-num-mid"; bar_col = "#D97706"
        else:            badge_cls = "qstat-num-ok";  bar_col = "#059669"

        txt_clean = _strip_html(s["texte"])
        preview = txt_clean[:90] + ("…" if len(txt_clean) > 90 else "")

        st.markdown(f"""
        <div class="qstat-card">
          <div class="qstat-header">
            <span class="qstat-num {badge_cls}">#{i+1} · {taux:.0f}% erreurs</span>
            <span class="qstat-txt">{preview}</span>
          </div>
          <div class="qstat-bar-wrap">
            <div class="qstat-bar-bg" style="flex:1">
              <div class="qstat-bar-fill" style="width:{min(taux,100):.0f}%;background:{bar_col}"></div>
            </div>
            <span style="font-size:.8rem;font-weight:700;color:{bar_col};min-width:38px">{taux:.0f}%</span>
          </div>
          <div class="qstat-meta">
            <span class="qstat-pill pill-tot">📊 {total} réponse(s)</span>
            <span class="qstat-pill pill-ok">✅ {bonnes} correcte(s)</span>
            <span class="qstat-pill pill-err">❌ {mauvaises} erreur(s)</span>
            <span class="qstat-pill" style="background:#F1F5F9;color:#475569">📝 {s.get('quiz_titre','')}</span>
          </div>
        </div>
        """, unsafe_allow_html=True)

    # ── Export Excel ──
    st.markdown("---")
    slbl("Export")
    import io as _io
    df_export = pd.DataFrame([{
        "Quiz":             s.get("quiz_titre", ""),
        "Code":             s.get("quiz_code", ""),
        "Question":         _strip_html(s["texte"]),
        "Type":             s["type"],
        "Points":           s["points"],
        "Total réponses":   int(s["total_reponses"] or 0),
        "Bonnes réponses":  int(s["bonnes_reponses"] or 0),
        "Erreurs":          int(s["mauvaises_reponses"] or 0),
        "Taux d'erreur (%)": float(s["taux_erreur"] or 0),
    } for s in stats])

    buf = _io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name="Stats questions")
        ws = writer.sheets["Stats questions"]
        # Mise en forme : largeur colonnes
        for col in ws.columns:
            ml = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 4, 60)
        # Coloration conditionnelle de la colonne "Taux d'erreur"
        from openpyxl.styles import PatternFill, Font
        taux_col_idx = df_export.columns.get_loc("Taux d'erreur (%)") + 1
        for row in ws.iter_rows(min_row=2, min_col=taux_col_idx, max_col=taux_col_idx):
            for cell in row:
                v = cell.value or 0
                if v >= 50:   cell.fill = PatternFill("solid", fgColor="FEE2E2"); cell.font = Font(bold=True, color="991B1B")
                elif v >= 25: cell.fill = PatternFill("solid", fgColor="FEF3C7"); cell.font = Font(color="92400E")
                else:         cell.fill = PatternFill("solid", fgColor="D1FAE5"); cell.font = Font(color="065F46")
    buf.seek(0)
    st.download_button(
        "📥  Télécharger les stats questions (Excel)",
        data=buf.getvalue(),
        file_name=f"stats_questions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


# ══════════════════════════════════════════════════════════════
#  ADMIN — Éditeur Quiz
# ══════════════════════════════════════════════════════════════

_TYPES={"single":"⭕  Choix unique","multiple":"☑️  Choix multiple","numeric":"🔢  Valeur numérique","text":"📝  Texte libre"}


def render_admin_quiz_edit():
    if not st.session_state.admin_logged: go("admin_login"); return
    quiz_id=st.session_state.edit_quiz_id; is_new=quiz_id is None; existing=get_quiz(quiz_id) if not is_new else None
    topbar("Nouveau Quiz" if is_new else "Modifier le Quiz","📝")
    st.markdown('<div class="card">',unsafe_allow_html=True)
    with st.form("fqi"):
        titre=st.text_input("Titre *",value=existing["titre"] if existing else "",placeholder="Ex : Évaluation module Femme")
        c1,c2=st.columns(2)
        with c1: code=st.text_input("Code *",value=existing["code"] if existing else "",placeholder="QZ001",max_chars=20)
        with c2: duree=st.number_input("Durée (min)",min_value=1,max_value=600,value=existing["duree_minutes"] if existing else 30)
        desc=st.text_area("Description",value=existing["description"] if existing else "",height=68)
        c1,c2=st.columns(2)
        with c1: show_sc=st.checkbox("Afficher le score aux agents",value=bool(existing.get("show_score",1)) if existing else True,help="Décoché = l'agent voit seulement 'Réponses enregistrées'")
        with c2: rnd=st.checkbox("Questions aléatoires",value=bool(existing.get("randomize_questions",0)) if existing else False,help="Ordre différent pour chaque agent")

        st.markdown("---")
        st.markdown("**⚠️ Points négatifs**")
        c1,c2=st.columns(2)
        with c1: malus_actif=st.checkbox("Activer les malus",value=bool(existing.get("malus_actif",0)) if existing else False,help="Mauvaise réponse = perte de points. Rien coché = 0 pt.")
        with c2: malus_pts=st.number_input("Points perdus / mauvaise réponse",min_value=0.0,max_value=10.0,step=0.5,value=float(existing.get("malus_points",0) or 0) if existing else 0.0,disabled=not malus_actif)

        c1,c2=st.columns(2)
        with c1: show_corr=st.checkbox("Permettre aux agents de revoir leur corrigé",value=bool(existing.get("show_correction",0)) if existing else False,help="Un bouton apparaît sur la page résultat — sans afficher le score")
        with c2: st.markdown("")
        st.markdown("**🔒 Anti-quitter**")
        anticheat=st.checkbox("Soumettre automatiquement si l'agent quitte l'écran (3 sorties max)",value=bool(existing.get("anticheat_actif",0)) if existing else False)
        if anticheat:
            all_ag=get_all_agents()
            try: ac_list=json.loads(existing.get("anticheat_agents","[]") or "[]") if existing else []
            except Exception: ac_list=[]
            sel_ag=st.multiselect("Appliquer à (vide = tous les agents)",
                options=[a["id"] for a in all_ag],
                default=[x for x in ac_list if x in [a["id"] for a in all_ag]],
                format_func=lambda x: next((f"{a['nom']} {a['prenom']}" for a in all_ag if a["id"]==x),"?"),
                key="ac_agents_sel")
        else:
            sel_ag=[]

        if st.form_submit_button("💾  Enregistrer",type="primary",use_container_width=True):
            if not titre.strip() or not code.strip(): st.error("Titre et code obligatoires.")
            else:
                ac_json=json.dumps(sel_ag)
                if is_new:
                    try:
                        nid=create_quiz(titre.strip(),code.strip(),int(duree),desc.strip(),int(show_sc),int(rnd),int(malus_actif),float(malus_pts),int(anticheat),ac_json,int(show_corr))
                        st.session_state.edit_quiz_id=nid; st.success("Quiz créé !"); st.rerun()
                    except Exception as e: st.error(f"Erreur (code déjà existant ?) : {e}")
                else:
                    update_quiz(quiz_id,titre.strip(),code.strip(),int(duree),desc.strip(),existing["actif"],int(show_sc),int(rnd),int(malus_actif),float(malus_pts),int(anticheat),ac_json,int(show_corr)); st.success("Quiz mis à jour."); st.rerun()
    st.markdown('</div>',unsafe_allow_html=True)

    cur_id=st.session_state.edit_quiz_id
    if not cur_id:
        if st.button("← Retour",key="qe_back_top",use_container_width=True): go("admin_dashboard")
        return

    questions=get_questions(cur_id)
    slbl(f"Questions ({len(questions)})")
    if questions:
        for i,q in enumerate(questions):
            with st.expander(f"Q{i+1}  ·  {_TYPES[q['type']]}  ·  {q['texte'][:50]}{'…' if len(q['texte'])>50 else ''}"):
                if q["type"] in("single","multiple"):
                    for o in q["options"]: st.markdown(f"{'✅' if o['is_correct'] else '◻️'} {o['texte']}")
                elif q["type"]=="numeric": st.markdown(f"**Réponse :** `{q['reponse_correcte_num']}`")
                elif q["type"]=="text":    st.markdown(f"**Réponses :** `{q['reponse_correcte_txt']}`")
                st.caption(f"{q['points']} point(s)")
                if st.button("🗑️ Supprimer",key=f"dq_{q['id']}"): delete_question(q["id"]); reorder_questions(cur_id); st.rerun()
    else: st.info("Aucune question. Ajoutez-en ci-dessous.")

    slbl("Ajouter une question")

    # Sélecteur de type HORS form (contrôle l'affichage conditionnel)
    q_type = st.selectbox(
        "Type de question",
        list(_TYPES.keys()),
        format_func=lambda x: _TYPES[x],
        key="aqt",
        label_visibility="collapsed"
    )

    # ── Éditeur riche HORS form (les composants HTML ne peuvent pas être dans st.form) ──
    slbl("Texte de la question")
    rte_val = rich_text_editor("Question *", key="rte_new_q",
                               default_value=st.session_state.get("rte_new_q", ""))

    # ── Reste du formulaire ──
    with st.form("faq", clear_on_submit=True):
        q_pts = st.number_input("Points", min_value=0.5, max_value=20.0, value=1.0, step=0.5)

        opts_d = []
        if q_type in ("single", "multiple"):
            st.markdown("**Options** — cochez la/les bonne(s) réponse(s) ✅")
            for j in range(6):
                c1, c2 = st.columns([5, 1])
                with c1:
                    ot = st.text_input(f"Option {j+1}", key=f"ot_{j}",
                                       label_visibility="collapsed",
                                       placeholder=f"Option {j+1}…")
                with c2:
                    oc = st.checkbox("✅", key=f"oc_{j}")
                if ot.strip():
                    opts_d.append({"texte": ot.strip(), "is_correct": oc})

        c_num = None
        c_txt = ""
        if q_type == "numeric":
            c_num = st.number_input("Réponse correcte", value=0.0, format="%.4f")
        elif q_type == "text":
            c_txt = st.text_input("Réponse(s)",
                                  placeholder="rep1 | rep2  (laisser vide = question ouverte)")

        submitted = st.form_submit_button("➕  Ajouter la question",
                                          type="primary", use_container_width=True)
        if submitted:
            q_txt = (rte_val or "").strip()
            err = None
            if not q_txt:
                err = "Texte de la question obligatoire."
            elif q_type in ("single", "multiple"):
                if not opts_d:
                    err = "Ajoutez au moins une option."
                elif not any(o["is_correct"] for o in opts_d):
                    err = "Cochez au moins une bonne réponse."
                elif q_type == "single" and sum(o["is_correct"] for o in opts_d) > 1:
                    err = "Choix unique : une seule bonne réponse."
            if err:
                st.error(err)
            else:
                ordre = len(questions)
                if q_type in ("single", "multiple"):
                    add_question(cur_id, q_txt, q_type, ordre, q_pts, options=opts_d)
                elif q_type == "numeric":
                    add_question(cur_id, q_txt, q_type, ordre, q_pts, num=c_num)
                elif q_type == "text":
                    add_question(cur_id, q_txt, q_type, ordre, q_pts,
                                 txt=(c_txt.strip() if c_txt else ""))
                # Réinitialiser l'éditeur riche
                st.session_state["rte_new_q"] = ""
                st.success("Question ajoutée ✓")
                st.rerun()


    st.markdown("---")
    if st.button("← Retour au dashboard",key="qe_back_bot",use_container_width=True): go("admin_dashboard")


# ══════════════════════════════════════════════════════════════

def _tab_surveillance():
    quizzes=get_all_quizzes()
    if not quizzes: st.info("Aucun quiz."); return
    opts={0:"— Tous —"}; opts.update({q["id"]:f"{q['titre']} ({q['code']})" for q in quizzes})
    sel=st.selectbox(" ",list(opts.keys()),format_func=lambda x:opts[x],label_visibility="collapsed",key="surv_sel")
    data=get_surveillance(sel if sel else None)
    if not data: st.info("Aucune soumission."); return

    suspects     = [r for r in data if r.get("suspects")]
    app_susp     = len([r for r in data if len(r.get("unique_devices") or [])>1])
    quit_susp    = len([r for r in data if int(r.get("quit_count") or 0)>=3])
    vitesse_susp = len([r for r in data if "trop rapide" in " ".join(r.get("suspects") or []) or "réponse en" in " ".join(r.get("suspects") or [])])

    # ── KPI ──
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("📋 Soumissions",     len(data))
    c2.metric("⚠️ Cas suspects",    len(suspects),  delta=f"{len(suspects)} à vérifier" if suspects else None, delta_color="inverse")
    c3.metric("📱 Appareils susp.", app_susp,       delta="changement d'appareil" if app_susp else None, delta_color="inverse")
    c4.metric("⚡ Vitesse susp.",   vitesse_susp,   delta="trop rapide" if vitesse_susp else None, delta_color="inverse")

    # ── Résumé distribution scores ──
    if data:
        scores = [r["score_pct"] for r in data]
        avg_s  = round(sum(scores)/len(scores),1)
        above80= sum(1 for s in scores if s>=80)
        below30= sum(1 for s in scores if s<30)
        st.markdown(
            f'<div style="background:#F8FAFF;border:1.5px solid #E2E8F4;border-radius:10px;padding:11px 16px;margin:10px 0;font-size:.82rem;color:#475569">' +
            f'📈 Score moyen : <b style="color:#1D4ED8">{avg_s}%</b> &nbsp;·&nbsp; ' +
            f'🏆 ≥80% : <b style="color:#059669">{above80}</b> &nbsp;·&nbsp; ' +
            f'📉 <30% : <b style="color:#DC2626">{below30}</b> &nbsp;·&nbsp; ' +
            f'🚪 Quitté ≥3x : <b style="color:#D97706">{quit_susp}</b></div>',
            unsafe_allow_html=True)

    st.markdown("---")
    only_s=st.checkbox("Cas suspects uniquement",key="surv_filter")
    rows=suspects if only_s else data
    if only_s and not rows: st.success("✅ Aucun cas suspect détecté."); return

    for r in rows:
        is_s   = bool(r.get("suspects"))
        border = "#DC2626" if is_s else "#E2E8F4"
        bg     = "#FFF5F5" if is_s else "#FFFFFF"
        pct    = r.get("score_pct",0)
        sc     = "#059669" if pct>=70 else("#D97706" if pct>=50 else "#DC2626")
        init   = (r["nom"][0]+(r["prenom"][0] if r.get("prenom") else "")).upper()
        duree  = r.get("duree_affichee","—")
        quit_c = int(r.get("quit_count") or 0)
        tps_m  = r.get("tps_moyen_rep")
        tps_mn = r.get("tps_min_rep")

        # Badges suspects
        badges=""
        for s in (r.get("suspects") or []):
            if "appareils" in s:    badges+=f'<span style="background:#FEE2E2;color:#DC2626;font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">📱 {s}</span>'
            elif "rapide" in s:     badges+=f'<span style="background:#FEF3C7;color:#D97706;font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">⚡ {s}</span>'
            elif "quitté" in s:     badges+=f'<span style="background:#FEF3C7;color:#D97706;font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">🚪 {s}</span>'
            elif "score élevé" in s:badges+=f'<span style="background:#FEE2E2;color:#DC2626;font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">🎯 {s}</span>'
            elif "réponse en" in s: badges+=f'<span style="background:#FEF3C7;color:#D97706;font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">⚡ {s}</span>'
            elif "sessions" in s:   badges+=f'<span style="background:#FEE2E2;color:#DC2626;font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">⚠️ {s}</span>'

        # Badge quit (si pas déjà dans suspects)
        if quit_c>0 and not any("quitté" in s for s in (r.get("suspects") or [])):
            qcol="#DC2626" if quit_c>=3 else "#D97706"
            badges+=f'<span style="background:#FFF7ED;color:{qcol};font-size:.7rem;font-weight:800;padding:2px 8px;border-radius:99px;margin-right:4px">🚪 Quitté {quit_c}x</span>'

        # Appareils
        unique_devs=r.get("unique_devices") or []
        if len(unique_devs)>1:
            devs_str="".join(f'<div style="color:#64748B;font-size:.72rem;margin-top:2px">• {d[:55]}</div>' for d in unique_devs)
            dev_html=f'<div style="background:#FEE2E2;border:1px solid #FCA5A5;border-radius:8px;padding:7px 10px;margin-top:6px"><b style="color:#DC2626;font-size:.78rem">⚠️ {len(unique_devs)} appareils différents ont composé cette session</b>{devs_str}</div>'
        elif r.get("device_info"):
            dev_html=f'<span style="background:#F1F5F9;color:#475569;font-size:.7rem;padding:2px 8px;border-radius:99px">📱 {r["device_info"][:40]}</span>'
        else:
            dev_html=""

        # Timing
        tps_html=""
        if tps_m is not None:
            tps_fmt = f"{int(tps_m)//60}m{int(tps_m)%60:02d}s" if tps_m>=60 else f"{tps_m}s"
            tps_mn_fmt = f"{tps_mn}s" if tps_mn is not None else "—"
            tcol = "#DC2626" if (tps_mn is not None and tps_mn<3) else "#7C3AED"
            tps_html=f'<span style="background:#EDE9FE;color:{tcol};font-size:.7rem;font-weight:700;padding:2px 8px;border-radius:99px;margin-left:4px">⏱ Moy:{tps_fmt} Min:{tps_mn_fmt}</span>'

        ip_html=f'<span style="background:#F1F5F9;color:#475569;font-size:.7rem;padding:2px 8px;border-radius:99px;margin-left:4px">🌐 {r["ip_address"]}</span>' if r.get("ip_address") else ""

        # Barre de durée
        dm2 = r.get("duree_minutes",30)*60
        ds2 = r.get("duree_secondes")
        pct_t = min(100,round(ds2/dm2*100)) if ds2 and dm2 else 0
        bar_col = "#059669" if pct_t>40 else("#D97706" if pct_t>20 else "#DC2626")
        dur_bar = f'<div style="background:#F1F5F9;border-radius:99px;height:5px;margin:5px 0 2px"><div style="width:{pct_t}%;height:5px;background:{bar_col};border-radius:99px"></div></div>' if ds2 else ""

        st.markdown(
            f'<div style="background:{bg};border:1.5px solid {border};border-radius:12px;padding:14px 16px;margin:6px 0">'+
            f'<div style="display:flex;align-items:center;gap:12px;margin-bottom:7px">'+
            f'<div style="width:36px;height:36px;border-radius:50%;background:#DBEAFE;color:#1D4ED8;display:flex;align-items:center;justify-content:center;font-weight:800;font-size:13px;flex-shrink:0">{init}</div>'+
            f'<div style="flex:1"><div style="font-weight:700;font-size:14px">{r["nom"]} {r.get("prenom","")}'+
            (f' <span style="color:#94A3B8;font-size:.78rem">({r["matricule"]})</span>' if r.get("matricule") else "")+
            f'</div><div style="font-size:.77rem;color:#64748B">{r.get("quiz_titre","")}</div></div>'+
            f'<div style="text-align:right;flex-shrink:0"><div style="font-size:1.05rem;font-weight:900;color:{sc}">{pct}%</div>'+
            f'<div style="font-size:.7rem;color:#94A3B8">{duree}</div></div></div>'+
            f'{dur_bar}'+
            f'<div style="display:flex;flex-wrap:wrap;gap:4px;align-items:center;margin-top:4px">{badges}{tps_html}{ip_html}</div>'+
            f'{dev_html}</div>',
            unsafe_allow_html=True)

    # ── Export Excel ──
    st.markdown("---")
    df=pd.DataFrame([{
        "Agent":          f"{r['nom']} {r.get('prenom','')}",
        "Matricule":      r.get("matricule",""),
        "Quiz":           r.get("quiz_titre",""),
        "Score (%)":      r.get("score_pct",0),
        "Durée":          r.get("duree_affichee","—"),
        "Nb appareils":   len(r.get("unique_devices") or []),
        "Appareils":      " | ".join(r.get("unique_devices") or []),
        "IP":             r.get("ip_address","—"),
        "Quitté (x)":     int(r.get("quit_count") or 0),
        "Tps moy/rép":    f"{r['tps_moyen_rep']}s" if r.get("tps_moyen_rep") else "—",
        "Tps min/rép":    f"{r['tps_min_rep']}s"   if r.get("tps_min_rep")   else "—",
        "Suspect":        ", ".join(r.get("suspects") or []) or "Non",
        "Soumis le":      r.get("completed_at",""),
    } for r in rows])
    buf=io.BytesIO()
    with pd.ExcelWriter(buf,engine="openpyxl") as w:
        df.to_excel(w,index=False,sheet_name="Surveillance")
        ws=w.sheets["Surveillance"]
        from openpyxl.styles import PatternFill, Font
        for row in ws.iter_rows(min_row=2):
            susp=row[11].value
            if susp and susp!="Non":
                for cell in row: cell.fill=PatternFill("solid",fgColor="FEE2E2")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width=min(max(len(str(c.value or ""))+4 for c in col),50)
    buf.seek(0)
    st.download_button("📥 Exporter rapport Excel",data=buf.getvalue(),
        file_name=f"surveillance_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,key="dl_surv")


def render_agent_correction():
    """Page de révision du devoir — sans afficher le score."""
    agent = st.session_state.current_agent
    quiz  = st.session_state.current_quiz
    sid   = st.session_state.get("result_session_id") or st.session_state.get("session_id")

    if not agent or not quiz or not sid:
        go("home"); return

    topbar(f"Corrigé — {quiz['titre']}", "📖")

    st.markdown(
        f'<div style="background:#EFF6FF;border:1.5px solid #BFDBFE;border-radius:12px;padding:12px 16px;margin:10px 0">' +
        f'<b style="color:#1D4ED8">📋 {quiz["titre"]}</b> &nbsp;·&nbsp; ' +
        f'<span style="color:#475569">{agent["nom"]} {agent.get("prenom","")}</span></div>',
        unsafe_allow_html=True)

    answers = get_session_answers(sid)
    if not answers:
        st.info("Aucune réponse enregistrée pour cette session.")
        if st.button("← Retour",key="corr_back_top",use_container_width=True): go("agent_result")
        return

    total_q  = len(answers)
    correctes = sum(1 for a in answers if a["is_correct"])

    # Résumé visuel sans score numérique
    pct = round(correctes/total_q*100) if total_q else 0
    bar_col = "#059669" if pct>=70 else ("#D97706" if pct>=50 else "#DC2626")
    st.markdown(
        f'<div style="background:#fff;border:1.5px solid #E2E8F4;border-radius:12px;padding:14px 16px;margin:8px 0">' +
        f'<div style="font-size:.82rem;color:#64748B;font-weight:600;margin-bottom:6px">{correctes} bonne(s) réponse(s) sur {total_q}</div>' +
        f'<div style="background:#F1F5F9;border-radius:99px;height:8px;overflow:hidden">' +
        f'<div style="width:{pct}%;height:8px;background:{bar_col};border-radius:99px"></div></div></div>',
        unsafe_allow_html=True)

    TYPES = {"single":"Choix unique","multiple":"Choix multiple","numeric":"Numérique","text":"Texte libre"}

    for i, a in enumerate(answers):
        ok      = bool(a["is_correct"])
        bg      = "#F0FDF4" if ok else "#FFF5F5"
        border  = "#6EE7B7" if ok else "#FCA5A5"
        ico     = "✅" if ok else "❌"
        pts_txt = f"+{a['points']:.0f}pt" if ok else "0pt"
        pts_col = "#059669" if ok else "#DC2626"

        # Construire la réponse agent lisible
        resp_raw = a.get("reponse","")
        try:
            import json as _j
            resp_parsed = _j.loads(resp_raw)
        except Exception:
            resp_parsed = resp_raw

        agent_rep = ""
        correct_rep = ""

        if a["type"] in ("single","multiple"):
            opts = a.get("options",[])
            # Réponse agent
            if isinstance(resp_parsed, list):
                rep_ids = resp_parsed
            elif resp_parsed:
                rep_ids = [resp_parsed]
            else:
                rep_ids = []
            agent_txts   = [o["texte"] for o in opts if o["id"] in rep_ids]
            correct_txts = [o["texte"] for o in opts if o["is_correct"]]
            agent_rep   = " / ".join(agent_txts) if agent_txts else "— Pas de réponse —"
            correct_rep = " / ".join(correct_txts)
        elif a["type"] == "numeric":
            agent_rep   = str(resp_raw) if str(resp_raw).strip() else "— Pas de réponse —"
            correct_rep = str(a.get("reponse_correcte_num",""))
        elif a["type"] == "text":
            agent_rep   = str(resp_raw) if str(resp_raw).strip() else "— Pas de réponse —"
            raw_cor     = (a.get("reponse_correcte_txt") or "").strip()
            correct_rep = raw_cor if raw_cor else "Question ouverte"

        st.markdown(
            f'<div style="background:{bg};border:1.5px solid {border};border-left:4px solid {border};border-radius:0 12px 12px 0;padding:14px 16px;margin:10px 0">' +
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">' +
            f'<span style="background:#1D4ED8;color:#fff;font-size:.7rem;font-weight:900;padding:2px 8px;border-radius:99px">Q{i+1}</span>' +
            f'<span style="color:#94A3B8;font-size:.75rem">{TYPES.get(a["type"],"")}</span>' +
            f'<span style="margin-left:auto;font-size:.8rem;font-weight:800;color:{pts_col}">{ico} {pts_txt}</span></div>' +
            f'<p style="font-size:15px;font-weight:600;color:#0F172A;margin:0 0 10px;line-height:1.5">{a["texte"]}</p>' +
            f'<div style="background:rgba(0,0,0,.04);border-radius:8px;padding:8px 12px;margin-bottom:6px;font-size:.85rem">' +
            f'<span style="color:#64748B;font-weight:600">Votre réponse : </span>' +
            f'<span style="color:#0F172A;font-weight:700">{agent_rep}</span></div>' +
            (f'<div style="background:rgba(5,150,105,.08);border-radius:8px;padding:8px 12px;font-size:.85rem">' +
             f'<span style="color:#059669;font-weight:600">Bonne réponse : </span>' +
             f'<span style="color:#065F46;font-weight:700">{correct_rep}</span></div>'
             if not ok and a["type"]!="text" or (a["type"]=="text" and (a.get("reponse_correcte_txt") or "").strip()) else "") +
            f'</div>',
            unsafe_allow_html=True)

    st.markdown("---")
    if st.button("← Retour aux résultats",key="corr_back",use_container_width=True): go("agent_result")

#  ROUTEUR
# ══════════════════════════════════════════════════════════════

_R={"home":render_home,"agent_search":render_agent_search,"agent_quiz_code":render_agent_quiz_code,
    "agent_quiz":render_agent_quiz,"agent_result":render_agent_result,"agent_correction":render_agent_correction,
    "admin_login":render_admin_login,"admin_dashboard":render_admin_dashboard,
    "admin_quiz_edit":render_admin_quiz_edit}

fn=_R.get(st.session_state.get("page","home"))
if fn: fn()
else: go("home")
